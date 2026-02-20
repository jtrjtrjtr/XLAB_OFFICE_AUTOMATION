#!/usr/bin/env python3
"""
XLAB Pricing Calculator v4
- On-site production: hourly rates, default 9h/day, persons×9 in quantity
- Tech progression baked into effective_days
- Popis (client descriptions) from ceník
- Discount only from Technika
- Agency fee default 0

Usage: python3 calculate.py --cenik cenik.xlsx --input input.json --output result.json
"""
import argparse, json, sys, os

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '--break-system-packages', '-q'])
    import openpyxl

EUR_DIVISOR = 25
TECH_PROGRESSION = {1: 1.0, 2: 1.5, 3: 2.0, 5: 3.0}
DEFAULT_ONSITE_HOURS = 9  # standard day = 9 hours

TIER_MAP = {
    'nakladova': 'cost', 'nákladová': 'cost',
    'snizena': 'snizena', 'snížená': 'snizena', 'partner': 'snizena',
    'standard': 'standard', 'premium': 'premium',
}
TIER_DISPLAY = {'cost': 'Nákladová', 'snizena': 'Snížená / Partner', 'standard': 'Standard', 'premium': 'Premium'}

def r50(v):
    if v is None or v == 0: return 0
    return round(v/50)*50 if abs(v) >= 200 else round(v)

def load_cenik(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    data = {'lide': {}, 'technika': {}, 'produkce': {}}
    # Ceník v5: A=Položka, B=Popis, C=Jednotka, D=Nákladová, E=Snížená, F=Standard, G=Premium, H=Poznámka
    for sk, sn in [('lide','Lidé'),('technika','Technika'),('produkce','Produkce & Logistika')]:
        ws = wb[sn]
        cat = None
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
            name = row[0].value
            popis = row[1].value
            unit = row[2].value
            if not name: continue
            if not unit and not row[3].value: cat = name; continue
            if name == 'Media server' and not row[3].value: continue
            data[sk][name] = {
                'category': cat, 'unit': unit,
                'cost': row[3].value, 'snizena': row[4].value,
                'standard': row[5].value, 'premium': row[6].value,
                'popis': popis,
                'note': row[7].value,
            }
    wb.close()
    return data

def get_tech_factor(days):
    if days <= 1: return 1.0
    for b in sorted(TECH_PROGRESSION.keys()):
        if days <= b: return TECH_PROGRESSION[b]
    return TECH_PROGRESSION[max(TECH_PROGRESSION.keys())]

def lookup(cenik, name, sheet):
    src = cenik.get(sheet, {})
    if name in src: return src[name]
    for k, v in src.items():
        if k.lower() == name.lower(): return v
    for k, v in src.items():
        if name.lower() in k.lower() or k.lower() in name.lower(): return v
    return None

# On-site production categories (matched case-insensitive)
ONSITE_CATEGORIES = ['on-site produkce', 'on-site', 'onsite produkce', 'onsite']

def is_onsite(category):
    return category.lower().strip() in ONSITE_CATEGORIES

def calculate(cenik, input_data):
    meta = input_data['metadata']
    tier_key = TIER_MAP.get(meta.get('tier','standard').lower(), 'standard')
    tier_display = TIER_DISPLAY.get(tier_key, 'Standard')
    agency_rate = meta.get('agency_fee_rate', 0.0)
    discount_rate = meta.get('discount_rate', 0.0)

    items = input_data.get('items', [])
    result_items, warnings = [], []
    cat_totals_sell = {}
    cat_totals_cost = {}
    technika_subtotal = 0

    for item in items:
        name = item['name']
        sheet = item.get('sheet', 'lide')
        qty = item.get('quantity', 1)
        days = item.get('days', 1)
        unit = item.get('unit', '')
        category = item.get('category', 'Ostatní')
        is_daily_tech = item.get('is_daily_tech', False)
        custom_price = item.get('custom_price', None)
        # On-site: persons field (optional, defaults to quantity for backward compat)
        persons = item.get('persons', None)

        pd = lookup(cenik, name, sheet)
        if not pd and not custom_price:
            warnings.append(f"'{name}' nenalezena ({sheet})")
            continue

        if custom_price:
            up_sell, up_cost = custom_price, round(custom_price * 0.6)
            popis = item.get('popis', '')
        else:
            up_sell = pd.get(tier_key, pd.get('standard', 0)) or 0
            up_cost = pd.get('cost', 0) or 0
            popis = item.get('popis') or pd.get('popis', '') or ''

        # ── On-site production: convert persons → hours ──
        # Rates in ceník are hourly. Default 9h/day.
        # Input can specify persons=N → output quantity = N × 9, unit = "h"
        # Or use_daily=true to skip conversion (manual daily rate override)
        use_daily = item.get('use_daily', False)

        if is_onsite(category) and not use_daily:
            hours_per_day = item.get('hours_per_day', DEFAULT_ONSITE_HOURS)
            if persons is not None:
                # Explicit persons: quantity = persons × hours
                qty = persons * hours_per_day
            elif unit in ('os/den', 'os/h', 'osob'):
                # Legacy: treat quantity as persons count
                qty = qty * hours_per_day
            # else: quantity already in hours (manual)
            unit = 'h'

        # Tech progression: bake into effective days
        eff_days = days
        if is_daily_tech and days > 1:
            eff_days = get_tech_factor(days)

        total_sell = round(qty * eff_days * up_sell)
        total_cost = round(qty * eff_days * up_cost)
        margin = round((1 - total_cost/total_sell)*100, 1) if total_sell > 0 else 0

        result_items.append({
            'name': name, 'category': category, 'sheet': sheet,
            'quantity': qty, 'days': eff_days,
            'unit': unit or (pd.get('unit','') if pd else ''),
            'unit_price_sell': up_sell, 'unit_price_cost': up_cost,
            'total_sell': total_sell, 'total_cost': total_cost,
            'margin_pct': margin, 'popis': popis,
        })

        cat_totals_sell[category] = cat_totals_sell.get(category, 0) + total_sell
        cat_totals_cost[category] = cat_totals_cost.get(category, 0) + total_cost
        if sheet == 'technika':
            technika_subtotal += total_sell

    subtotal_sell = sum(cat_totals_sell.values())
    subtotal_cost = sum(cat_totals_cost.values())

    discount = round(technika_subtotal * discount_rate)
    agency_fee = round((subtotal_sell - discount) * agency_rate)
    total_czk = subtotal_sell - discount + agency_fee
    total_eur = round(total_czk / EUR_DIVISOR)
    total_profit = total_czk - subtotal_cost
    overall_margin = round((1 - subtotal_cost/total_czk)*100, 1) if total_czk > 0 else 0

    return {
        'metadata': {**meta, 'tier_display': tier_display, 'agency_fee_rate': agency_rate, 'discount_rate': discount_rate},
        'items': result_items,
        'category_totals': {k: {'sell': cat_totals_sell.get(k,0), 'cost': cat_totals_cost.get(k,0)} for k in cat_totals_sell},
        'technika_subtotal': technika_subtotal,
        'summary': {
            'subtotal_sell': subtotal_sell, 'subtotal_cost': subtotal_cost,
            'technika_subtotal': technika_subtotal,
            'discount': discount, 'discount_rate': discount_rate,
            'agency_fee': agency_fee, 'agency_fee_rate': agency_rate,
            'total_czk': total_czk, 'total_eur': total_eur,
            'total_cost': subtotal_cost, 'total_profit': total_profit,
            'overall_margin_pct': overall_margin,
        },
        'warnings': warnings,
    }

def find_cenik():
    """Auto-find ceník: user upload > skill data > ask user."""
    candidates = [
        '/mnt/user-data/uploads/XLAB_Centralni_Cenik_v5.xlsx',
        '/mnt/user-data/uploads/XLAB_Centralni_Cenik_v4.xlsx',
    ]
    # Skill data directory (bundled ceník)
    script_dir = os.path.dirname(os.path.abspath(__file__))
    skill_data = os.path.join(os.path.dirname(script_dir), 'data')
    import glob
    for f in sorted(glob.glob(os.path.join(skill_data, 'XLAB_Centralni_Cenik_v*.xlsx')), reverse=True):
        candidates.append(f)
    # Also check common locations
    candidates.extend([
        '/mnt/skills/user/xlab-pricing/data/XLAB_Centralni_Cenik_v5.xlsx',
    ])
    for path in candidates:
        if os.path.exists(path):
            return path
    return None


def calculate_freeform(input_data):
    """Freeform calculation — no ceník lookup, all prices from input."""
    meta = input_data['metadata']
    agency_rate = meta.get('agency_fee_rate', 0.0)
    discount_rate = meta.get('discount_rate', 0.0)

    items = input_data.get('items', [])
    result_items, warnings = [], []
    cat_totals_sell = {}
    cat_totals_cost = {}
    technika_subtotal = 0

    for item in items:
        name = item['name']
        category = item.get('category', 'Ostatní')
        qty = item.get('quantity', 1)
        days = item.get('days', 1)
        unit = item.get('unit', '')
        up_sell = item.get('unit_price_sell', item.get('custom_price', 0))
        up_cost = item.get('unit_price_cost', round(up_sell * 0.6))
        popis = item.get('popis', '')
        is_daily_tech = item.get('is_daily_tech', False)

        eff_days = days
        if is_daily_tech and days > 1:
            eff_days = get_tech_factor(days)

        total_sell = round(qty * eff_days * up_sell)
        total_cost = round(qty * eff_days * up_cost)
        margin = round((1 - total_cost / total_sell) * 100, 1) if total_sell > 0 else 0

        result_items.append({
            'name': name, 'category': category, 'sheet': item.get('sheet', 'custom'),
            'quantity': qty, 'days': eff_days, 'unit': unit,
            'unit_price_sell': up_sell, 'unit_price_cost': up_cost,
            'total_sell': total_sell, 'total_cost': total_cost,
            'margin_pct': margin, 'popis': popis,
        })

        cat_totals_sell[category] = cat_totals_sell.get(category, 0) + total_sell
        cat_totals_cost[category] = cat_totals_cost.get(category, 0) + total_cost
        if item.get('sheet') == 'technika' or item.get('is_technika', False):
            technika_subtotal += total_sell

    subtotal_sell = sum(cat_totals_sell.values())
    subtotal_cost = sum(cat_totals_cost.values())
    discount = round(technika_subtotal * discount_rate)
    agency_fee = round((subtotal_sell - discount) * agency_rate)
    total_czk = subtotal_sell - discount + agency_fee
    total_eur = round(total_czk / EUR_DIVISOR)
    total_profit = total_czk - subtotal_cost
    overall_margin = round((1 - subtotal_cost / total_czk) * 100, 1) if total_czk > 0 else 0

    return {
        'metadata': {**meta, 'tier_display': meta.get('tier', 'Custom'), 'agency_fee_rate': agency_rate, 'discount_rate': discount_rate},
        'items': result_items,
        'category_totals': {k: {'sell': cat_totals_sell.get(k, 0), 'cost': cat_totals_cost.get(k, 0)} for k in cat_totals_sell},
        'technika_subtotal': technika_subtotal,
        'summary': {
            'subtotal_sell': subtotal_sell, 'subtotal_cost': subtotal_cost,
            'technika_subtotal': technika_subtotal,
            'discount': discount, 'discount_rate': discount_rate,
            'agency_fee': agency_fee, 'agency_fee_rate': agency_rate,
            'total_czk': total_czk, 'total_eur': total_eur,
            'total_cost': subtotal_cost, 'total_profit': total_profit,
            'overall_margin_pct': overall_margin,
        },
        'warnings': warnings,
    }


def main():
    p = argparse.ArgumentParser()
    p.add_argument('--cenik', default=None, help='Path to ceník XLSX (auto-detected if omitted)')
    p.add_argument('--input', required=True)
    p.add_argument('--output', required=True)
    p.add_argument('--freeform', action='store_true', help='Freeform mode — no ceník, all prices from input')
    a = p.parse_args()

    with open(a.input, 'r', encoding='utf-8') as f: inp = json.load(f)

    # Check for freeform flag in input or CLI
    is_freeform = a.freeform or inp.get('metadata', {}).get('freeform', False)

    if is_freeform:
        result = calculate_freeform(inp)
        print("Režim: FREEFORM (bez ceníku)")
    else:
        cenik_path = a.cenik or find_cenik()
        if not cenik_path:
            print("CHYBA: Ceník nenalezen. Nahraj XLAB_Centralni_Cenik_v5.xlsx nebo použij --freeform")
            sys.exit(1)
        print(f"Ceník: {cenik_path}")
        cenik = load_cenik(cenik_path)
        result = calculate(cenik, inp)
    with open(a.output, 'w', encoding='utf-8') as f: json.dump(result, f, ensure_ascii=False, indent=2)
    s = result['summary']
    print(f"Kalkulace: {len(result['items'])} položek")
    print(f"  Subtotal:   {s['subtotal_sell']:>10,} CZK  (technika: {s['technika_subtotal']:,})")
    if s['discount']: print(f"  Sleva HW:   {-s['discount']:>10,} CZK")
    if s['agency_fee']: print(f"  Agency fee: {s['agency_fee']:>10,} CZK ({s['agency_fee_rate']*100:.1f}%)")
    print(f"  CELKEM:     {s['total_czk']:>10,} CZK / {s['total_eur']:,} EUR")
    print(f"  Náklady:    {s['total_cost']:>10,} CZK")
    print(f"  Zisk:       {s['total_profit']:>10,} CZK  |  Marže: {s['overall_margin_pct']:.1f}%")
    for w in result.get('warnings',[]): print(f"  ⚠ {w}")

if __name__ == '__main__': main()
