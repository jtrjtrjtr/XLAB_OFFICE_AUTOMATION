#!/usr/bin/env python3
"""
XLAB Pricing Excel Generator v6
- Grayscale palette with warm gray-yellow accents
- Internal section stays orange
- Summary sheet for PPTX export
- File naming: CE_JobNr_Klient_Projekt_v01.xlsx
- Footer: IČO, DIČ, BANKA
- On-site: hourly rates (9h default)
- Design: Popis header, left-aligned B/C, v-center all items,
  D column no fill, unified cat/total fill

Usage: python3 generate_excel.py --input result.json --output Kalkulace.xlsx [--assets /path/to/assets]
"""
import argparse, json, sys, os, re
from datetime import datetime
from collections import OrderedDict

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.drawing.image import Image as XlImage
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '--break-system-packages', '-q'])
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.drawing.image import Image as XlImage

# ═══════════════════════════════════════════════
# PALETTE — grayscale with warm gray-yellow tones
# ═══════════════════════════════════════════════
CHARCOAL  = '333333'   # primary text, dark headers
MID       = '666666'   # secondary text
LIGHT_TXT = '888888'   # notes, subtle text
BLACK     = '000000'   # footer text
WHITE     = 'FFFFFF'

# Warm grays (gray mixed with yellow/beige)
WARM_1    = 'F7F3EC'   # lightest — editable cell accent (E, G columns)
WARM_2    = 'EDE7DC'   # internal category header fills (K-O)
WARM_3    = 'E3DDD2'   # subtotal fills
WARM_BIG  = 'BFB8AD'   # category headers A-H + CELKEM row A-H (unified)
WARM_4    = 'D4CEC3'   # internal CELKEM fill (K-O)

# Internal section — stays orange
ORANGE    = 'CC6600'
GREEN     = '2E7D32'
INT_EDIT  = 'FFF8E7'   # light yellow for internal editable cells

# Navy kept only for title "Cenová kalkulace"
NAVY      = '1F4E79'

# ═══════════════════════════════════════════════
# FONTS — all Helvetica Light
# ═══════════════════════════════════════════════
def F(sz=10, bold=False, italic=False, color=CHARCOAL, name='Helvetica Light'):
    return Font(name=name, size=sz, bold=bold, italic=italic, color=color)

F_TITLE    = F(16, bold=True, color=NAVY)
F_LABEL    = F(10, bold=True, color=CHARCOAL)
F_VALUE    = F(10, color=CHARCOAL)
F_HDR      = F(10, bold=True, color=WHITE)
F_HDR_L    = F(10, bold=True, color=WHITE)   # left-aligned header variant
F_NORMAL   = F(10, color=CHARCOAL)
F_NOTE     = F(9, italic=True, color=LIGHT_TXT)
F_CAT      = F(10, bold=True, color=CHARCOAL)
F_SUB_L    = F(10, bold=True, color=CHARCOAL)
F_SUB_V    = F(10, bold=True, color=CHARCOAL)
F_TOT_L    = F(12, bold=True, color=CHARCOAL)
F_TOT_V    = F(12, bold=True, color=CHARCOAL)
F_EDIT     = F(10, bold=True, color=CHARCOAL)
F_MZS_L    = F(10, bold=True, color=CHARCOAL)
F_MZS_V    = F(10, bold=True, color=CHARCOAL)
F_DISC_L   = F(10, bold=True, color=CHARCOAL)
# Internal
F_COST     = F(10, color=ORANGE)
F_COST_B   = F(10, bold=True, color=ORANGE)
F_MARGIN   = F(9, color=ORANGE)
F_MARGIN_B = F(10, bold=True, color=ORANGE)
F_PROFIT   = F(10, bold=True, color=GREEN)
F_FOOTER   = F(8, color=BLACK)
F_FOOTER9  = F(9, color=BLACK)
F_SUMHDR   = F(10, bold=True, color=WHITE)
F_SUMCAT   = F(10, color=CHARCOAL)
F_SUMTOT   = F(10, bold=True, color=CHARCOAL)

# ═══════════════════════════════════════════════
# FILLS
# ═══════════════════════════════════════════════
FILL_HDR     = PatternFill('solid', fgColor=CHARCOAL)
FILL_INT     = PatternFill('solid', fgColor=ORANGE)
FILL_CAT     = PatternFill('solid', fgColor=WARM_BIG)   # category headers A-H
FILL_CAT_INT = PatternFill('solid', fgColor=WARM_2)     # category headers K-O (lighter)
FILL_SUB     = PatternFill('solid', fgColor=WARM_3)
FILL_TOTAL   = PatternFill('solid', fgColor=WARM_BIG)   # CELKEM A-H (same as category)
FILL_TOT_INT = PatternFill('solid', fgColor=WARM_4)     # CELKEM K-O
FILL_WARM1   = PatternFill('solid', fgColor=WARM_1)     # editable accent (E, G only)
FILL_IEDIT   = PatternFill('solid', fgColor=INT_EDIT)   # internal editable

# Alignments
NFMT = '#,##0'; PFMT = '0.0%'; TXTFMT = '@'; DATEFMT = '[$]dd\\.mm\\.yyyy;@'
CENTER   = Alignment(horizontal='center')
CENTER_V = Alignment(horizontal='center', vertical='center')
RIGHT_A  = Alignment(horizontal='right')
LEFT_A   = Alignment(horizontal='left')
LEFT_V   = Alignment(horizontal='left', vertical='center')
VCENTER  = Alignment(vertical='center')
VCENTER_LW = Alignment(horizontal='left', vertical='center', wrap_text=True)


def generate_filename(meta):
    """Generate filename: CE_JobNr_Klient_Projekt_v01.xlsx"""
    job = meta.get('job_nr', 'JobNr')
    client = meta.get('client', 'Klient')
    project = meta.get('event', 'Projekt')
    def clean(s):
        return re.sub(r'[^\w\s-]', '', str(s)).strip().replace(' ', '_')[:30]
    return f"CE_{clean(job)}_{clean(client)}_{clean(project)}_v01.xlsx"


def generate_excel(result, output_path, assets_dir=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Kalkulace'

    # Column widths
    for letter, w in {'A':3, 'B':31.3, 'C':19.5, 'D':11, 'E':7, 'F':9, 'G':5, 'H':11.5,
                      'I':1.5, 'J':13, 'K':7, 'L':5, 'M':11, 'N':13, 'O':8}.items():
        ws.column_dimensions[letter].width = w

    meta = result['metadata']
    items = result['items']
    summary = result['summary']

    def cell(row, col, val, font=F_NORMAL, align=None, fmt=None, fill=None):
        c = ws.cell(row=row, column=col, value=val)
        c.font = font
        if align: c.alignment = align
        if fmt: c.number_format = fmt
        if fill: c.fill = fill
        return c

    def fill_row_range(row, fill, col_start, col_end):
        for col in range(col_start, col_end + 1):
            ws.cell(row=row, column=col).fill = fill

    # ═══════════════════════════════════════
    # HEADER — R1 logo space, R2-R5 meta
    # ═══════════════════════════════════════
    if assets_dir:
        logo_path = os.path.join(assets_dir, 'xlab_wordmark.png')
        if os.path.exists(logo_path):
            img = XlImage(logo_path)
            img.width = 180; img.height = 32
            ws.add_image(img, 'A2')

    # R2: Klient
    cell(2, 5, 'Klient:', F_LABEL, RIGHT_A)
    ws.merge_cells('F2:H2')
    cell(2, 6, meta.get('client', ''), F_VALUE)

    # R3: Projekt + Tier
    cell(3, 5, 'Projekt', F_LABEL, RIGHT_A)
    ws.merge_cells('F3:H3')
    cell(3, 6, meta.get('event', ''), F_VALUE)
    cell(3, 11, 'Tier', F_LABEL)
    cell(3, 12, meta.get('tier_display', 'Standard'), F_VALUE)

    # R4: Job Nr.
    cell(4, 5, 'Job Nr.', F_LABEL, RIGHT_A)

    # R5: Title + Datum
    ws.merge_cells('A5:B5')
    cell(5, 1, 'Cenová kalkulace', F_TITLE)
    ws.row_dimensions[5].height = 22
    cell(5, 5, 'Datum:', F_LABEL, RIGHT_A)
    ws.merge_cells('F5:H5')
    date_str = meta.get('date', '')
    try:
        from datetime import datetime as dt
        date_val = dt.strptime(date_str, '%Y-%m-%d') if date_str else dt.now()
    except:
        date_val = datetime.now()
    cell(5, 6, date_val, F_VALUE, LEFT_A, fmt=DATEFMT)

    # ═══════════════════════════════════════
    # R7: TABLE HEADERS
    # ═══════════════════════════════════════
    r = 7
    cell(r, 1, '', font=F(10, color=CHARCOAL), fill=FILL_HDR)
    # B and C: LEFT aligned
    cell(r, 2, 'Položka', F_HDR, LEFT_A, fill=FILL_HDR)
    cell(r, 3, 'Popis', F_HDR, LEFT_A, fill=FILL_HDR)
    for col, txt in [(4,'Cena/jed.'), (5,'Mn.'), (6,'Jednotka'), (7,'Dní'), (8,'Celkem CZK')]:
        cell(r, col, txt, F_HDR, CENTER, fill=FILL_HDR)
    for col, txt in [(11,'Mn.'), (12,'Dní'), (13,'Nákl./jed.'), (14,'Nákl. celk.'), (15,'Marže')]:
        cell(r, col, txt, F_HDR, CENTER, fill=FILL_INT)

    # Tip text → Q7-Q8 (outside print)
    ws.merge_cells(f'Q{r}:X{r}')
    cell(r, 17, '💡 Žlutě podbarvené buňky = editovatelné. Celky se přepočítají automaticky.', F_NOTE)
    ws.merge_cells(f'Q{r+1}:X{r+1}')
    cell(r+1, 17, f"Generováno: {datetime.now().strftime('%d.%m.%Y %H:%M')}  |  XLAB s.r.o.  |  www.xlab.cz", F_NOTE)

    r = 8

    # ═══════════════════════════════════════
    # ITEMS — grouped by category
    # ═══════════════════════════════════════
    categories = OrderedDict()
    for item in items:
        categories.setdefault(item.get('category', 'Ostatní'), []).append(item)

    cat_sub_rows = {}
    technika_sub_rows = []
    all_sub_rows = []

    for cat, cat_items in categories.items():
        # Category header — WARM_BIG for A-H, WARM_2 for K-O
        ws.merge_cells(f'A{r}:H{r}')
        cell(r, 1, cat.upper(), F_CAT, fill=FILL_CAT)
        fill_row_range(r, FILL_CAT_INT, 11, 15)
        r += 1

        item_rows = []
        for item in cat_items:
            popis = item.get('popis', '')

            # B: name — always v-center
            cell(r, 2, item['name'], F_NORMAL, VCENTER)

            # C: popis
            if popis:
                cell(r, 3, popis, F_NOTE, VCENTER_LW)
                if len(popis) > 40:
                    ws.row_dimensions[r].height = 40

            # D: unit price — NO fill, v-center
            cell(r, 4, item['unit_price_sell'], F_NORMAL, VCENTER, fmt=NFMT)
            # E: quantity — WARM_1 fill, center+v-center
            cell(r, 5, item['quantity'], F_NORMAL, CENTER_V, fill=FILL_WARM1)
            # F: unit — center+v-center, no fill
            cell(r, 6, item.get('unit', ''), F_NORMAL, CENTER_V)
            # G: days — WARM_1 fill, center+v-center
            days_val = item['days']
            c_d = cell(r, 7, days_val, F_NORMAL, CENTER_V, fill=FILL_WARM1)
            if isinstance(days_val, float) and days_val != int(days_val):
                c_d.number_format = '0.0'
            # H: total — v-center
            cell(r, 8, f'=D{r}*E{r}*G{r}', F_NORMAL, VCENTER, fmt=NFMT)

            # Internal columns
            cell(r, 11, item['quantity'], F_COST, CENTER, fill=FILL_IEDIT)
            c_id = cell(r, 12, days_val, F_COST, CENTER, fill=FILL_IEDIT)
            if isinstance(days_val, float) and days_val != int(days_val):
                c_id.number_format = '0.0'
            cell(r, 13, item['unit_price_cost'], F_COST, fmt=NFMT, fill=FILL_IEDIT)
            cell(r, 14, f'=M{r}*K{r}*L{r}', F_COST, fmt=NFMT)
            cell(r, 15, f'=IF(H{r}=0,0,1-N{r}/H{r})', F_MARGIN, CENTER, fmt=PFMT)

            item_rows.append(r)
            r += 1

        # Subtotal — "Celkem <category>"
        sub_r = r
        ws.merge_cells(f'A{r}:G{r}')
        cell(r, 1, f'Celkem {cat}', F_SUB_L, RIGHT_A)
        if item_rows:
            cell(r, 8, f'=SUM(H{item_rows[0]}:H{item_rows[-1]})', F_SUB_V, fmt=NFMT)
            cell(r, 14, f'=SUM(N{item_rows[0]}:N{item_rows[-1]})', F_COST_B, fmt=NFMT)
        fill_row_range(r, FILL_SUB, 1, 15)

        cat_sub_rows[cat] = sub_r
        all_sub_rows.append(sub_r)
        if any(it.get('sheet') == 'technika' for it in cat_items):
            technika_sub_rows.append(sub_r)
        r += 1

    r += 1  # spacer

    # ═══════════════════════════════════════
    # SUMMARY SECTION
    # ═══════════════════════════════════════
    grand_sub_row = r
    ws.merge_cells(f'A{r}:G{r}')
    cell(r, 1, 'MEZISOUČET', F_MZS_L, RIGHT_A)
    sub_refs = '+'.join([f'H{sr}' for sr in all_sub_rows])
    cell(r, 8, f'={sub_refs}', F_MZS_V, fmt=NFMT)
    cost_refs = '+'.join([f'N{sr}' for sr in all_sub_rows])
    cell(r, 14, f'={cost_refs}', F_COST_B, fmt=NFMT)
    r += 1

    # Sleva z techniky
    disc_row = r
    ws.merge_cells(f'A{r}:E{r}')
    cell(r, 1, 'Sleva z techniky', F_DISC_L, RIGHT_A)
    cell(r, 6, summary.get('discount_rate', 0.0), F_EDIT, CENTER, fmt=PFMT, fill=FILL_WARM1)
    if technika_sub_rows:
        hw_refs = '+'.join([f'H{sr}' for sr in technika_sub_rows])
        cell(r, 8, f'=-({hw_refs})*F{disc_row}', F_MZS_V, fmt=NFMT)
    else:
        cell(r, 8, 0, F_MZS_V, fmt=NFMT)
    r += 1

    # Agency fee
    fee_row = r
    ws.merge_cells(f'A{r}:E{r}')
    cell(r, 1, 'Agency fee', F_DISC_L, RIGHT_A)
    ws.row_dimensions[r].height = 16
    cell(r, 6, summary.get('agency_fee_rate', 0.0), F_EDIT, CENTER, fmt=PFMT, fill=FILL_WARM1)
    cell(r, 8, f'=(H{grand_sub_row}+H{disc_row})*F{fee_row}', F_MZS_V, fmt=NFMT)
    r += 1

    # CELKEM bez DPH — WARM_BIG fill for A-H (same as categories), WARM_4 for K-O
    total_row = r
    ws.merge_cells(f'A{r}:G{r}')
    ws.row_dimensions[r].height = 17
    cell(r, 1, 'CELKEM bez DPH', F_TOT_L, LEFT_A, fill=FILL_TOTAL)
    cell(r, 8, f'=H{grand_sub_row}+H{disc_row}+H{fee_row}', F_TOT_V, fmt=NFMT, fill=FILL_TOTAL)
    fill_row_range(r, FILL_TOTAL, 1, 8)       # A-H: WARM_BIG
    fill_row_range(r, FILL_TOT_INT, 11, 15)   # K-O: WARM_4
    r += 3

    # ═══════════════════════════════════════
    # FOOTER
    # ═══════════════════════════════════════
    if assets_dir:
        icon_path = os.path.join(assets_dir, 'xlab_x_icon.png')
        if os.path.exists(icon_path):
            img2 = XlImage(icon_path)
            img2.width = 50; img2.height = 31
            ws.add_image(img2, f'A{r}')

    # Row 1: M + phone | IČO | XLAB s.r.o. | INTERNÍ SOUHRN
    cell(r, 3, 'M', F_FOOTER, RIGHT_A)
    cell(r, 4, ' +420 273 136 947', F_FOOTER, fmt=TXTFMT)
    cell(r, 5, 'IČO', F_FOOTER, RIGHT_A)
    cell(r, 6, '020 56 623', F_FOOTER, LEFT_A)
    cell(r, 8, 'XLAB s.r.o.', F_FOOTER, LEFT_A)
    ws.merge_cells(f'K{r}:O{r}')
    cell(r, 11, 'INTERNÍ SOUHRN', F_COST_B)
    r += 1

    # Row 2: E + email | DIČ | Výstaviště 67 | Celkové náklady
    cell(r, 3, 'E', F_FOOTER, RIGHT_A)
    cell(r, 4, 'xlab@xlab.cz', F_FOOTER9)
    cell(r, 5, 'DIČ', F_FOOTER, RIGHT_A)
    cell(r, 6, 'CZ02056623', F_FOOTER, LEFT_A)
    cell(r, 8, 'Výstaviště 67', F_FOOTER, LEFT_A)
    ws.merge_cells(f'K{r}:M{r}')
    cell(r, 11, 'Celkové náklady', F_COST, RIGHT_A)
    cell(r, 14, f'=N{grand_sub_row}', F_COST_B, fmt=NFMT)
    r += 1

    # Row 3: W + url | BANKA | Praha 7 | Celkový zisk
    cell(r, 3, 'W', F_FOOTER, RIGHT_A)
    cell(r, 4, 'www.xlab.cz', F_FOOTER9)
    cell(r, 5, 'BANKA', F_FOOTER, RIGHT_A)
    cell(r, 6, '260540361/0300', F_FOOTER, LEFT_A)
    cell(r, 8, '170 00 Praha 7', F_FOOTER, LEFT_A)
    ws.merge_cells(f'K{r}:M{r}')
    cell(r, 11, 'Celkový zisk', F_PROFIT, RIGHT_A)
    cell(r, 14, f'=H{total_row}-N{grand_sub_row}', F_PROFIT, fmt=NFMT)
    r += 1

    # Celková marže
    ws.row_dimensions[r].height = 16
    ws.merge_cells(f'K{r}:M{r}')
    cell(r, 11, 'Celková marže', F_MARGIN_B, RIGHT_A)
    cell(r, 14, f'=IF(H{total_row}=0,0,1-N{grand_sub_row}/H{total_row})', F_MARGIN_B, fmt=PFMT)

    # Print setup
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.orientation = 'landscape'
    ws.print_area = f'A1:H{total_row + 6}'

    # ═══════════════════════════════════════
    # SUMMARY SHEET
    # ═══════════════════════════════════════
    ws2 = wb.create_sheet('Summary')
    ws2.column_dimensions['A'].width = 3
    ws2.column_dimensions['B'].width = 35
    ws2.column_dimensions['C'].width = 15
    ws2.column_dimensions['D'].width = 3

    def cell2(row, col, val, font=F_NORMAL, align=None, fmt=None, fill=None):
        c = ws2.cell(row=row, column=col, value=val)
        c.font = font
        if align: c.alignment = align
        if fmt: c.number_format = fmt
        if fill: c.fill = fill
        return c

    # Title
    ws2.merge_cells('A2:C2')
    cell2(2, 1, 'Cenová kalkulace — souhrn', F_TITLE)
    ws2.row_dimensions[2].height = 21

    # Meta
    cell2(3, 2, f"{meta.get('client', '')} — {meta.get('event', '')}", F_VALUE)

    # Table header
    sr = 5
    cell2(sr, 1, '', fill=FILL_HDR)
    cell2(sr, 2, 'Kategorie', F_HDR, fill=FILL_HDR)
    cell2(sr, 3, 'Celkem CZK', F_HDR, CENTER, fill=FILL_HDR)
    sr += 1

    # Categories — alternating warm accent
    for i, cat in enumerate(categories.keys()):
        sub_r = cat_sub_rows[cat]
        row_fill = FILL_WARM1 if i % 2 == 0 else None
        if row_fill:
            cell2(sr, 1, '', fill=row_fill)
            cell2(sr, 2, cat, F_SUMCAT, fill=row_fill)
            cell2(sr, 3, f"=Kalkulace!H{sub_r}", F_SUMCAT, fmt=NFMT, fill=row_fill)
        else:
            cell2(sr, 2, cat, F_SUMCAT)
            cell2(sr, 3, f"=Kalkulace!H{sub_r}", F_SUMCAT, fmt=NFMT)
        sr += 1

    sr += 1  # spacer

    # CELKEM bez DPH
    tot_fill = PatternFill('solid', fgColor=WARM_4)
    cell2(sr, 1, '', fill=tot_fill)
    cell2(sr, 2, 'CELKEM bez DPH', F_SUMTOT, fill=tot_fill)
    cell2(sr, 3, f"=Kalkulace!H{total_row}", F_SUMTOT, fmt=NFMT, fill=tot_fill)

    ws2.print_area = f'A1:C{sr + 2}'

    wb.save(output_path)
    print(f"Excel uložen: {output_path}")


def main():
    p = argparse.ArgumentParser()
    p.add_argument('--input', required=True)
    p.add_argument('--output', required=True)
    p.add_argument('--assets', default=None)
    a = p.parse_args()
    with open(a.input, 'r', encoding='utf-8') as f: result = json.load(f)
    assets = a.assets
    if not assets:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        candidate = os.path.join(os.path.dirname(script_dir), 'assets')
        if os.path.exists(candidate): assets = candidate

    out_path = a.output
    if os.path.isdir(out_path):
        out_path = os.path.join(out_path, generate_filename(result['metadata']))

    generate_excel(result, out_path, assets)

if __name__ == '__main__': main()
